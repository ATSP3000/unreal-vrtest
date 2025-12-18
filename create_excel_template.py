#!/usr/bin/env python3
"""
Creates the Excel template for the roadmap data input.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_template():
    wb = Workbook()
    
    # ========== Settings Sheet ==========
    settings = wb.active
    settings.title = "Settings"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A1A4E")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    settings["A1"] = "Setting"
    settings["B1"] = "Value"
    settings["A1"].font = header_font
    settings["B1"].font = header_font
    settings["A1"].fill = header_fill
    settings["B1"].fill = header_fill
    
    settings_data = [
        ("Title", "Dummy data Infrastructure Capstone Roadmap"),
        ("Navy Color", "1A1A4E"),
        ("Pink Background", "FFE0E0"),
        ("Purple Background", "E0D8F0"),
        ("Yellow Background", "FFF0D0"),
        ("Milestone Color", "F0C040"),
        ("Critical Text Color", "CC0000"),
        ("Near Term Color", "555555"),
        ("Mid Term Color", "C8A080"),
        ("Far Term Color", "E8D8C8"),
    ]
    
    for i, (setting, value) in enumerate(settings_data, start=2):
        settings[f"A{i}"] = setting
        settings[f"B{i}"] = value
    
    settings.column_dimensions["A"].width = 20
    settings.column_dimensions["B"].width = 50
    
    # ========== Timeline Sheet ==========
    timeline = wb.create_sheet("Timeline")
    
    timeline["A1"] = "Year"
    timeline["B1"] = "Width (inches)"
    timeline["C1"] = "Is Last Column"
    for col in ["A", "B", "C"]:
        timeline[f"{col}1"].font = header_font
        timeline[f"{col}1"].fill = header_fill
    
    years_data = [
        ("2023", 0.8, "No"),
        ("2024", 0.8, "No"),
        ("2025", 0.8, "No"),
        ("2026", 0.8, "No"),
        ("2027", 0.8, "No"),
        ("2028", 0.8, "No"),
        ("2029-2053", 1.8, "Yes"),
    ]
    
    for i, (year, width, is_last) in enumerate(years_data, start=2):
        timeline[f"A{i}"] = year
        timeline[f"B{i}"] = width
        timeline[f"C{i}"] = is_last
    
    timeline.column_dimensions["A"].width = 15
    timeline.column_dimensions["B"].width = 15
    timeline.column_dimensions["C"].width = 15
    
    # ========== Goals Sheet ==========
    goals = wb.create_sheet("Goals")
    
    goals["A1"] = "Goal ID"
    goals["B1"] = "Goal Name"
    for col in ["A", "B"]:
        goals[f"{col}1"].font = header_font
        goals[f"{col}1"].fill = header_fill
    
    goals_data = [
        (1, "Overarching goal 1"),
        (2, "Overarching goal 2"),
    ]
    
    for i, (goal_id, name) in enumerate(goals_data, start=2):
        goals[f"A{i}"] = goal_id
        goals[f"B{i}"] = name
    
    goals.column_dimensions["A"].width = 10
    goals.column_dimensions["B"].width = 30
    
    # ========== Rows Sheet ==========
    rows = wb.create_sheet("Rows")
    
    row_headers = ["Row ID", "STC Label", "FTA Label", "Background Color", "Goal ID", "Row Height (inches)"]
    for col_idx, header in enumerate(row_headers, start=1):
        cell = rows.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Row data: (row_id, stc, fta, bg_color, goal_id, height)
    rows_data = [
        (1, "STC1", "FTA1", "Pink", 1, 0.6),
        (2, "STC1", "FTA2", "Pink", 1, 0.6),
        (3, "STC3", "FTA3", "Purple", 1, 0.6),
        (4, "STC2", "FTA4", "Purple", 1, 0.6),
        (5, "STC4", "FTA5", "Yellow", 2, 0.6),
        (6, "STC4", "FTA6", "Yellow", 2, 0.6),
    ]
    
    for i, row_data in enumerate(rows_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            rows.cell(row=i, column=col_idx, value=value)
    
    for col_idx in range(1, 7):
        rows.column_dimensions[get_column_letter(col_idx)].width = 18
    
    # ========== Milestones Sheet ==========
    milestones = wb.create_sheet("Milestones")
    
    milestone_headers = ["Row ID", "Year", "Vertical Offset", "Text", "Is Critical"]
    for col_idx, header in enumerate(milestone_headers, start=1):
        cell = milestones.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Milestones: (row_id, year, v_offset, text, is_critical)
    milestones_data = [
        # FTA1 milestones
        (1, 2024, 0.15, "Critical Goal Text", "Yes"),
        (1, 2027, 0.2, "Goal text", "No"),
        (1, 2028, 0.35, "Goal text", "No"),
        # FTA2 milestones
        (2, 2025, 0.1, "Goal text", "No"),
        (2, 2026, 0.08, "Critical Goal Text", "Yes"),
        (2, 2026, 0.25, "Critical Goal Text", "Yes"),
        (2, 2026, 0.42, "Critical Goal Text", "Yes"),
        (2, 2027, 0.1, "Goal text", "No"),
        (2, 2029, 0.25, "Critical Goal Text", "Yes"),
        # FTA3 milestones
        (3, 2025, 0.35, "Goal text", "No"),
        (3, 2026, 0.1, "Goal text", "No"),
        (3, 2027, 0.4, "Goal text", "No"),
        (3, 2028, 0.15, "Goal text", "No"),
        # FTA4 milestones
        (4, 2025, 0.12, "Goal text", "No"),
        (4, 2027, 0.22, "Critical Goal Text", "Yes"),
        (4, 2027, 0.42, "Goal text", "No"),
        # FTA5 milestones
        (5, 2024, 0.12, "Goal text", "No"),
        (5, 2026, 0.08, "Goal text", "No"),
        (5, 2026, 0.28, "Goal text", "No"),
        (5, 2027, 0.38, "Critical Goal Text", "Yes"),
        # FTA6 milestones
        (6, 2023, 0.2, "Goal text", "No"),
        (6, 2027, 0.08, "Critical Goal Text", "Yes"),
        (6, 2027, 0.25, "Goal text", "No"),
        (6, 2027, 0.42, "Critical Goal Text", "Yes"),
    ]
    
    for i, ms_data in enumerate(milestones_data, start=2):
        for col_idx, value in enumerate(ms_data, start=1):
            milestones.cell(row=i, column=col_idx, value=value)
    
    milestones.column_dimensions["A"].width = 10
    milestones.column_dimensions["B"].width = 12
    milestones.column_dimensions["C"].width = 15
    milestones.column_dimensions["D"].width = 25
    milestones.column_dimensions["E"].width = 12
    
    # ========== Use Cases Sheet ==========
    usecases = wb.create_sheet("UseCases")
    
    uc_headers = ["Use Case ID", "Description", "Color"]
    for col_idx, header in enumerate(uc_headers, start=1):
        cell = usecases.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    uc_data = [
        ("UC1, UC2, UC3, UC4", "All use cases", "F0C040"),
        ("UC1, UC2", "Use case 1 text, Use case 2 text", "E8A060"),
        ("UC2, UC4", "Use case 2 text, Use case 4 text", "F0D080"),
    ]
    
    for i, data in enumerate(uc_data, start=2):
        for col_idx, value in enumerate(data, start=1):
            usecases.cell(row=i, column=col_idx, value=value)
    
    usecases.column_dimensions["A"].width = 20
    usecases.column_dimensions["B"].width = 40
    usecases.column_dimensions["C"].width = 12
    
    # ========== Instructions Sheet ==========
    instructions = wb.create_sheet("Instructions")
    
    instructions["A1"] = "ROADMAP DATA INPUT INSTRUCTIONS"
    instructions["A1"].font = Font(bold=True, size=14)
    
    inst_text = [
        "",
        "This Excel file is used to populate the Infrastructure Capstone Roadmap PowerPoint.",
        "",
        "SHEETS:",
        "",
        "1. Settings - Configure title and colors",
        "   - Title: The main title of the roadmap",
        "   - Colors: Hex color codes (without #) for various elements",
        "",
        "2. Timeline - Define the years shown in the header",
        "   - Year: The label to show (e.g., '2023' or '2029-2053')",
        "   - Width: Column width in inches",
        "   - Is Last Column: 'Yes' for the final column",
        "",
        "3. Goals - Define overarching goals",
        "   - Goal ID: Unique identifier (1, 2, etc.)",
        "   - Goal Name: Text to display",
        "",
        "4. Rows - Define swim lane rows",
        "   - Row ID: Unique identifier",
        "   - STC Label: Strategic Test Capability label",
        "   - FTA Label: Functional Test Area label",
        "   - Background Color: 'Pink', 'Purple', or 'Yellow'",
        "   - Goal ID: Which overarching goal this row belongs to",
        "   - Row Height: Height in inches",
        "",
        "5. Milestones - Define milestone markers",
        "   - Row ID: Which row this milestone belongs to",
        "   - Year: Which year column to place it in",
        "   - Vertical Offset: Position within row (0.0 to 0.5)",
        "   - Text: Label text",
        "   - Is Critical: 'Yes' for red critical text, 'No' for normal",
        "",
        "6. UseCases - Define legend use case indicators",
        "   - Use Case ID: Label for the use case combo",
        "   - Description: Description text",
        "   - Color: Hex color code for the triangle marker",
        "",
        "To generate the PowerPoint, run:",
        "   python create_roadmap_from_excel.py roadmap_data.xlsx",
    ]
    
    for i, text in enumerate(inst_text, start=2):
        instructions[f"A{i}"] = text
    
    instructions.column_dimensions["A"].width = 80
    
    # Move Instructions to first position
    wb.move_sheet(instructions, offset=-5)
    
    return wb


if __name__ == "__main__":
    wb = create_template()
    wb.save("roadmap_data.xlsx")
    print("Template saved to roadmap_data.xlsx")
